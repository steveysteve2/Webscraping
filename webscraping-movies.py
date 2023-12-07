
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font


#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2023/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

print(title.text)

table_rows = soup.findAll("tr")

# print(table_rows[:2])

wb = xl.Workbook()
ws = wb.active
ws.title = 'Box Office Report'

ws['A1'] = 'No.'
ws['B1'] = 'Movie Title'
ws['C1'] = 'Gross'
ws['D1'] = 'No. of Theaters'
ws['E1'] = 'Avg. Gross / Theater'

for x in range(1,6):
    #rank, title, gross, and theaters
    td = table_rows[x].findAll("td")
    rank = int(td[0].text)
    title = td[1].text
    gross = int((td[5].text).replace(",","").replace("$",""))
    theats = int((td[6].text).replace(",",""))
    avg = gross/theats

    #this evaluates to ws['A2'], ws['A3'], etc.
    ws['A' + str(x+1)] = rank
    ws['B' + str(x+1)] = title
    ws['C' + str(x+1)] = gross
    ws['D' + str(x+1)] = theats
    ws['E' + str(x+1)] = avg

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 25

header_font = Font(size=16,bold=True)

for cell in ws[1:1]:
    cell.font = header_font

for cell in ws["C:C"]:
    cell.number_format = u'"$"#,##0.00'

for cell in ws["D:D"]:
    cell.number_format = u'#,##0'

for cell in ws["E:E"]:
    cell.number_format = u'"$"#,##0.00'

wb.save('BoxOfficeReport.xlsx')

##
##
##
##

